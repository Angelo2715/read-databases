import pymysql
from platform import python_implementation
import openpyxl
import pandas as pd
# Give the location of the file
path = "agents.xlsx"
 
# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

agent_code=[]
agent_name=[]
working_area=[]
commision=[]
phone_no=[]
country=[]


sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
#declarar data frame
# Loop will print all values
# of first column
for i in range(1, m_row + 1):
    cell_obj1 = sheet_obj.cell(row = i, column = 1)
    cell_obj2 = sheet_obj.cell(row = i, column = 2)
    cell_obj3 = sheet_obj.cell(row = i, column = 3)
    cell_obj4 = sheet_obj.cell(row = i, column = 4)
    cell_obj5 = sheet_obj.cell(row = i, column = 5)
    cell_obj6 = sheet_obj.cell(row = i, column = 6)
    agent_code.append(cell_obj1.value)
    agent_name.append(cell_obj2.value)
    working_area.append(cell_obj3.value)
    commision.append(cell_obj4.value)
    phone_no.append(cell_obj5.value)
    country.append(cell_obj6.value)


############### CONFIGURAR ESTO ###################
# Abre conexion con la base de datos
db = pymysql.connect(host='localhost', user= 'angelo', passwd='', db='test')
##################################################

# prepare a cursor object using cursor() method
cursor = db.cursor()

# SENTENCIAS SQL PARA INSERTAR DATOS EN OTRAS TABLAS
#sql = "INSERT INTO test(id, name, email) \
#   VALUES ('{0}','{1}')".format("cosme","testmail@sever.com")
#-------------------------------otras secuencias sql para insertar datos en distintas tablas --------------------------------
#sql2="INSERT INTO company(COMPANY_ID, COMPANY_NAME, COMPANY_CITY ) \
#   VALUES ('{0}','{1}','{2}')".format("","")

#sql3="INSERT INTO customer(CUST_CODE, CUST_NAME, CUST_CITY, WORKING_AREA, CUST_COUNTRY, GRADE, OPENING_AMT, RECEIVE_AMT, PAYMENT_AMT, OUTSTANDING_AMT, PHONE_NO, AGENT_CODE) \
#   VALUES ('{0}','{1}','{2}','{3}','{4}','{5}','{6}','{7}','{8}','{9}','{10}','{11}')".format("","")
#
#sql4="INSERT INTO daysorder(ORD_NUM, ORD_AMOUNT, ADVANCE_AMOUNT, ORD_DATE, CUST_CODE, AGENT_CODE, ORD_DESCRIPTION ) \
#   VALUES ('{0}','{1}','{2}','{3}','{4}','{5}','{6}')".format("","")

#sql5="INSERT INTO despatch(DES_NUM, DES_DATE, DES_AMOUNT, ORD_NUM, ORD_DATE, ORD_AMOUNT, AGENT_CODE  ) \
#   VALUES ('{0}','{1}','{2}','{3}','{4}','{5}','{6}')".format("","")

#sql6="INSERT INTO foods(ITEM_ID, ITEM_NAME, ITEM_UNIT, COMPANY_ID ) \
#   VALUES ('{0}','{1}','{2}','{3}')".format("","")

#sql7="INSERT INTO listofitem(ITEMCODE, ITEMNAME, BATCHCODE, CONAME ) \
#   VALUES ('{0}','{1}','{2}','{3}')".format("","")

#sql7="INSERT INTO orders(ORD_NUM, ORD_AMOUNT, ORD_DATE, CUST_CODE, AGENT_CODE) \
#   VALUES ('{0}','{1}','{2}','{3}','{4}')".format("","")

#sql8="INSERT INTO student(NAME, TITLE, CLASS, SECTION, ROLLID) \
#   VALUES ('{0}','{1}','{2}','{3}','{4}')".format("","")

#sql8="INSERT INTO studentreport(CLASS, SECTION, ROLLID, GRADE, SEMISTER, CLASS_ATTENDED) \
#   VALUES ('{0}','{1}','{2}','{3}','{4}','{5}')".format("","")
#-------------------------------otras secuencias sql para insertar datos en distintas tablas --------------------------------  


for i in range(len(agent_code)):

   sql1="INSERT INTO agents(AGENT_CODE, AGENT_NAME, WORKING_AREA, COMMISSION, PHONE_NO, COUNTRY ) \
      VALUES ('{0}','{1}','{2}','{3}','{4}','{5}')".format(agent_code[i],agent_name[i],working_area[i],commision[i],phone_no[i],country[i])

   try:
      # Execute the SQL command
      cursor.execute(sql1)
      # Commit your changes in the database
      db.commit()
   except:
      # Rollback in case there is any error
      db.rollback()


# desconectar del servidor
db.close()